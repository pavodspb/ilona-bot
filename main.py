import openpyxl
from datetime import datetime, timedelta
from pathlib import Path
import os
import random
from groq import Groq

# Настройки — абсолютные пути
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "темы.xlsx")
POSTS_DIR = os.path.join(BASE_DIR, "посты")
os.makedirs(POSTS_DIR, exist_ok=True)

# Groq API
GROQ_KEY_FILE = os.path.join(BASE_DIR, "Илона.txt")
_client = None

def get_client():
    global _client
    if _client is not None:
        return _client
    api_key = ""
    if os.path.exists(GROQ_KEY_FILE):
        with open(GROQ_KEY_FILE, "r") as f:
            api_key = f.read().strip()
    if api_key:
        _client = Groq(api_key=api_key)
        return _client
    return None

def generate_with_ai(topic, platform):
    client = get_client()
    if not client:
        return None
    
    if platform == "vk":
        prompt = f"""Напиши пост для ВКонтакте на тему: "{topic}"

Требования:
- Объём: 100-200 слов
- Стиль: живой, неформальный, как будто пишет друг
- Начни с цепляющей фразы-хука
- Дай 3-4 конкретных совета
- Закончи вопросом к аудитории
- Добавь 2-3 хештега в конце
- Без воды и шаблонных фраз
- Пиши от первого лица"""
    else:
        prompt = f"""Напиши статью для Дзена на тему: "{topic}"

Требования:
- Объём: 300-500 слов
- Стиль: экспертный, но доступный
- Заголовок должен цеплять
- Структура: введение, основная часть с советами, вывод
- Дай 4-5 конкретных советов с примерами
- Закончи призывом к действию
- Без воды и шаблонов"""
    
    try:
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": "Ты — опытный маркетолог-копирайтер. Пиши живым языком, без канцелярита."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.9,
            max_tokens=1000
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Ошибка Groq: {e}")
        return None

def create_excel_if_not_exists():
    """Создает Excel-файл с темами, если он не существует"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Темы"
        ws.append(["Дата", "Тема", "Статус"])
        
        # Примеры тем
        sample_topics = [
            ["24.06.2026", "Как увеличить продажи через Instagram"],
            ["25.06.2026", "5 ошибок в таргетированной рекламе"],
            ["26.06.2026", "Секреты продающих текстов"],
            ["27.06.2026", "Email-рассылка: как не попасть в спам"],
            ["28.06.2026", "Анализ конкурентов: пошаговая инструкция"],
        ]
        
        for topic in sample_topics:
            ws.append([topic[0], topic[1], ""])
        
        # Форматирование
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        wb.save(EXCEL_FILE)
        print(f"Создан файл {EXCEL_FILE} с примерами тем")

def get_next_topic():
    """Находит следующую ненаписанную тему"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_str, topic, status = row
        if status != "написано" and topic:
            return date_str, topic
    
    return None, None

def mark_topic_as_done(date_str):
    """Отмечает тему как написанную"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2):
        if row[0].value == date_str:
            row[2].value = "написано"
            break
    
    wb.save(EXCEL_FILE)

def generate_post_template(topic):
    """Генерирует пост на заданную тему"""
    
    # Определяем категорию темы
    topic_lower = topic.lower()
    
    # Шаблоны в зависимости от типа темы
    if "кухн" in topic_lower or "мебел" in topic_lower or "купить" in topic_lower:
        hooks = [
            "Слушайте, а вы знали, что можно сэкономить кучу денег на кухне?",
            "Был уверен, что кухня — это дорого. Пока не узнал об этом.",
            "Кто-нибудь покупал кухню в последнее время? Делюсь находкой!",
            "Затарился кухней и решил поделиться лайфхаком для тех, кто ищет."
        ]
        intro = random.choice(hooks)
        body = f"{topic} — тема, которая меня зацепила. Оказалось, всё не так просто, как кажется на первый взгляд."
        tips = [
            "Обойдите минимум 3-4 магазина перед покупкой",
            "Спросите про «скрытые» скидки — они почти всегда есть",
            "Не ведитесь на «акцию последнего дня» — это маркетинг",
            "Проверьте реальные отзывы, а не те, что на сайте"
        ]
        cta = "А у вас был опыт? Как сэкономили (или потеряли)? 👇"
        
    elif "реклам" in topic_lower or "таргет" in topic_lower or "продвижени" in topic_lower:
        hooks = [
            "Слили бюджет на рекламу и не получили клиентов? Знакомо?",
            "Одна ошибка в таргете, и деньги улетают в трубу. Рассказываю, как не повторять.",
            "Знаете, почему чужая реклама работает, а ваша — нет?",
            "Братва, кто хоть раз запускал рекламу — поймёт боль."
        ]
        intro = random.choice(hooks)
        body = f"{topic} — тема, о которой молчат «гуру», но без этого никак. Разбираем по полочкам."
        tips = [
            "Начните с 500 рублей в день — не больше, пока не поймёте",
            "Тестируйте 3-4 варианта объявления одновременно",
            "Смотрите на клик-through, а не на охваты",
            "Отключайте то, что не работает — не жалейте"
        ]
        cta = "Какая ваша самая частая ошибка в рекламе? Делитесь 👇"
        
    elif "соцсет" in topic_lower or "instagram" in topic_lower or "вконтакт" in topic_lower or "вк " in topic_lower:
        hooks = [
            "Сидите в соцсетях часами, а толку — ноль? Давайте исправим.",
            "Завёл аккаунт для бизнеса, а подписчики молчат? Я тоже через это прошёл.",
            "Соцсети — это не просто «пост-картинка-репост». Вот что я имею в виду.",
            "Кто-нибудь замечал, что одни аккаунты растут, а другие — нет?"
        ]
        intro = random.choice(hooks)
        body = f"{topic} — тема, которая мне открыла глаза. Оказалось, я делал(а) всё не так."
        tips = [
            "Посты в 23:00 работают лучше, чем в 12:00 — проверено",
            "Отвечайте на каждый комментарий — даже негативный",
            "Видео даёт охват в 3-5 раз больше картинки",
            "Не пишите «продам» — пишите «помогу решить проблему»"
        ]
        cta = "А в какое время вы постите? Когда лучше всего заходит? 🤔"
        
    elif "цена" in topic_lower or "стоимост" in topic_lower or "дешевл" in topic_lower or "экономи" in topic_lower:
        hooks = [
            "Экономить — не значит жить плохо. Вот доказательство.",
            "Потратил 2 часа на сравнение цен и сэкономил 15 тысяч. Стоило?",
            "Все говорят «дешёвое — плохое». Это миф. Вот почему.",
            "Хотите сэкономить, но не знаете с чего начать? Начните с этого."
        ]
        intro = random.choice(hooks)
        body = f"{topic} — тема, которая касается каждого из нас. Разбираемся, где реальная экономия, а где — маркетинговый трюк."
        tips = [
            "Покупайте в сезон скидок — но только то, что реально нужно",
            "Не верьте цене «со скидкой» — проверяйте историю цен",
            "Кэшбэк и бонусы — это реальные деньги, пользуйтесь",
            "Один раз потратили время на сравнение — экономите каждый месяц"
        ]
        cta = "А какой у вас главный лайфхак по экономии? Делитесь секретами 👇"
        
    else:
        hooks = [
            "Слушайте, а вот это меня реально зацепило. Делюсь!",
            "Наткнулся на кое-что интересное. Решил рассказать.",
            "Бывает, находишь информацию и думаешь: «Почему я не знал раньше?!»",
            "Вот это — то, о чём很少 говорят, но все должны знать."
        ]
        intro = random.choice(hooks)
        body = f"{topic} — тема, которая заставила меня задуматься. Давайте разберёмся вместе."
        tips = [
            "Начните с малого — не пытайтесь объять всё сразу",
            "Ищите проверенные источники, а не «экспертов» из Instagram",
            "Делайте выводы из своего опыта, а не из чужих слов",
            "Записывайте — что сработало, а что нет"
        ]
        cta = "А что вы об этом думаете? Какой ваш опыт? 👇"

    # Формируем ВК пост
    vk_post = f"""{intro}

{topic}.

{body}

Мой топ-4 совета:

{chr(10).join(f'{i+1}. {tip}' for i, tip in enumerate(tips))}

{cta}

#маркетинг #продвижение #советы"""

    # Формируем Дзен пост
    dzen_post = f"""{topic}

{intro}

{body}

Что я выяснил(а) и что советую:

{chr(10).join(f'→ {tip}' for tip in tips)}

{cta}

#маркетинг #бизнес #советы"""

    return vk_post, dzen_post

def generate_post(topic):
    """Генерирует пост — сначала через DeepSeek, потом шаблон"""
    # Пробуем AI
    vk_ai = generate_with_ai(topic, "vk")
    dzen_ai = generate_with_ai(topic, "dzen")
    
    if vk_ai and dzen_ai:
        print("Сгенерировано через Groq")
        return vk_ai, dzen_ai
    
    # Фолбэк на шаблоны
    print("Groq недоступен, используем шаблоны")
    return generate_post_template(topic)

def save_post(date_str, topic, vk_post, dzen_post):
    """Сохраняет посты в файлы"""
    # Форматируем дату для имени файла
    try:
        date = datetime.strptime(date_str, "%d.%m.%Y")
        file_date = date.strftime("%Y-%m-%d")
    except:
        file_date = datetime.now().strftime("%Y-%m-%d")
    
    # Сохраняем ВК пост
    vk_file = os.path.join(POSTS_DIR, f"vk_{file_date}.txt")
    with open(vk_file, "w", encoding="utf-8") as f:
        f.write(vk_post)
    
    # Сохраняем Дзен пост
    dzen_file = os.path.join(POSTS_DIR, f"dzen_{file_date}.txt")
    with open(dzen_file, "w", encoding="utf-8") as f:
        f.write(dzen_post)
    
    print(f"Посты сохранены:")
    print(f"  ВК: {vk_file}")
    print(f"  Дзен: {dzen_file}")

def main():
    print("=== Маркетинговый помощник ===\n")
    
    # Создаем Excel если нет
    create_excel_if_not_exists()
    
    # Получаем следующую тему
    date_str, topic = get_next_topic()
    
    if not topic:
        print("Все темы уже использованы! Добавьте новые в темы.xlsx")
        return
    
    print(f"Тема на сегодня: {topic}")
    print(f"Дата: {date_str}\n")
    
    # Генерируем посты
    vk_post, dzen_post = generate_post(topic)
    
    # Сохраняем
    save_post(date_str, topic, vk_post, dzen_post)
    
    # Отмечаем как написанное
    mark_topic_as_done(date_str)
    print("\nТема отмечена как использованная.")

if __name__ == "__main__":
    main()
